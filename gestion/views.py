from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.contrib.auth.views import PasswordChangeView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.utils import timezone
from django.db.models import Sum, Prefetch, F, Count
from django.http import HttpResponse
from django.template.loader import render_to_string
from io import BytesIO
import datetime
import openpyxl
from openpyxl.styles import Font
from xhtml2pdf import pisa

from .models import (
    Proyecto, RegistroHoras, Actividad, Cliente,
    AsignacionProyecto, PerfilEmpleado,
)
from .forms import (
    ProyectoCreateForm, ProyectoUpdateForm, RegistroHorasForm,
    ClienteForm, EmpleadoForm, EmpleadoUpdateForm,
    CustomPasswordChangeForm, ReporteFiltroForm, AsignarProyectoForm,
)


# === LOGIN ===
def login_view(request):
    """Vista para iniciar sesiÃƒÂ³n."""
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)

            # Ã¢Å“â€¦ Registrar acciÃƒÂ³n en la bitÃƒÂ¡cora
            Actividad.objects.create(
                usuario=user,
                accion=f"IniciÃƒÂ³ sesiÃƒÂ³n en el sistema."
            )

            return redirect('admin_home' if user.is_staff else 'empleado_home')
    else:
        form = AuthenticationForm()
    return render(request, 'gestion/login.html', {'form': form})


# === LOGOUT ===
@login_required
def logout_view(request):
    """Cierra sesiÃƒÂ³n y redirige al login."""
    # Ã¢Å“â€¦ Registrar acciÃƒÂ³n en la bitÃƒÂ¡cora
    Actividad.objects.create(
        usuario=request.user,
        accion=f"CerrÃƒÂ³ sesiÃƒÂ³n."
    )

    logout(request)
    return redirect('login')


# === ADMINISTRADOR ===
@login_required
def admin_home(request):
    """Panel principal del administrador."""
    return render(request, 'gestion/admin_home.html')


# === EMPLEADO ===
@login_required
def empleado_home(request):
    """Panel principal del empleado."""
    return render(request, 'gestion/empleado_home.html')


# === GESTIÃƒâ€œN DE PROYECTOS (ADMIN) ===
@login_required
def lista_proyectos(request):
    """Muestra todos los proyectos (solo admin)."""
    if not request.user.is_staff:
        return redirect('empleado_home')
    proyectos = Proyecto.objects.select_related('cliente').all()
    return render(request, 'gestion/proyectos.html', {'proyectos': proyectos})


# AsegÃƒÂºrate de que esta lÃƒÂ­nea estÃƒÂ© al principio de tu archivo, junto a los otros imports
 

@login_required
def nuevo_proyecto(request):
    """Permite crear un nuevo proyecto (solo admin)."""
    if not request.user.is_staff:
        return redirect('empleado_home')

    if request.method == 'POST':
        # AquÃƒÂ­ usamos el nuevo formulario de creaciÃƒÂ³n
        form = ProyectoCreateForm(request.POST) 
        if form.is_valid():
            proyecto = form.save()

            # Ã¢Å“â€¦ Registrar acciÃƒÂ³n (si tienes el modelo Actividad)
            # Actividad.objects.create(
            #     usuario=request.user,
            #     accion=f"CreÃƒÂ³ el proyecto '{proyecto.nombre}'."
            # )

            return redirect('lista_proyectos')
    else:
        # AquÃƒÂ­ tambiÃƒÂ©n usamos el nuevo formulario de creaciÃƒÂ³n
        form = ProyectoCreateForm() 

    return render(request, 'gestion/nuevo_proyecto.html', {'form': form})


@login_required
def editar_proyecto(request, proyecto_id):
    """Permite al administrador editar un proyecto existente."""
    if not request.user.is_staff:
        return redirect('empleado_home')

    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        form = ProyectoUpdateForm(request.POST, instance=proyecto)
        if form.is_valid():
            cliente = form.save()
            Actividad.objects.create(usuario=request.user, accion="Registro el cliente '" + cliente.nombre + "'. ")

            # Ã¢Å“â€¦ Registrar acciÃƒÂ³n
            Actividad.objects.create(
                usuario=request.user,
                accion=f"EditÃƒÂ³ el proyecto '{proyecto.nombre}'."
            )

            return redirect('lista_proyectos')
    else:
        form = ProyectoUpdateForm(instance=proyecto)

    return render(request, 'gestion/editar_proyecto.html', {'form': form, 'proyecto': proyecto})


@login_required
def eliminar_proyecto(request, proyecto_id):
    """Permite al administrador eliminar un proyecto."""
    if not request.user.is_staff:
        return redirect('empleado_home')

    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        nombre = proyecto.nombre
        proyecto.delete()

        # Ã¢Å“â€¦ Registrar acciÃƒÂ³n
        Actividad.objects.create(
            usuario=request.user,
            accion=f"EliminÃƒÂ³ el proyecto '{nombre}'."
        )

        return redirect('lista_proyectos')

    return render(request, 'gestion/eliminar_proyecto.html', {'proyecto': proyecto})


# === REGISTRO DE HORAS (EMPLEADOS) ===
@login_required
def registrar_horas(request):
    """Permite al empleado registrar sus horas trabajadas."""
    if request.user.is_staff:
        return redirect('admin_home')

    if request.method == 'POST':
        # Ã°Å¸â€˜â€¡ *** AQUÃƒÂ: Pasamos el 'user' al formulario ***
        form = RegistroHorasForm(request.POST, user=request.user)
        
        if form.is_valid():
            registro = form.save(commit=False)
            registro.empleado = request.user
            registro.save()

            # Ã¢Å“â€¦ Registrar acciÃƒÂ³n
            Actividad.objects.create(
                usuario=request.user,
                accion=f"RegistrÃƒÂ³ {registro.horas} horas en el proyecto '{registro.proyecto.nombre}'."
            )

            return redirect('mis_horas')
    else:
        # Ã°Å¸â€˜â€¡ *** Y AQUÃƒÂ TAMBIÃƒâ€°N: Pasamos el 'user' al formulario vacÃƒÂ­o ***
        form = RegistroHorasForm(user=request.user)

    return render(request, 'gestion/registrar_horas.html', {'form': form})


@login_required
def mis_horas(request):
    """Muestra las horas registradas por el empleado autenticado."""
    if request.user.is_staff:
        return redirect('admin_home')

    horas = RegistroHoras.objects.filter(empleado=request.user).select_related('proyecto')
    return render(request, 'gestion/mis_horas.html', {'horas': horas})


# === ADMINISTRADOR: VER REGISTROS DE HORAS ===
@login_required
def ver_registros_horas_admin(request):
    """
    Muestra todos los registros de horas con filtros y resÃƒÂºmenes.
    Solo accesible para administradores.
    """
    if not request.user.is_staff:
        return redirect('empleado_home')

    # ===== PASO 1: DEFINE 'registros' Y FILTROS =====
    # Define la variable 'registros' aquÃƒÂ­ al principio
    registros = RegistroHoras.objects.select_related('empleado', 'proyecto')

    # Obtiene los IDs de los filtros ANTES de calcular resÃƒÂºmenes
    empleado_id = request.GET.get('empleado')
    proyecto_id = request.GET.get('proyecto')

    # Aplica los filtros a 'registros' si existen
    if empleado_id:
        registros = registros.filter(empleado_id=empleado_id)
    if proyecto_id:
        registros = registros.filter(proyecto_id=proyecto_id)
    # ===============================================

    # ===== PASO 2: CALCULA RESÃƒÅ¡MENES (usando 'registros') =====
    total_horas = registros.aggregate(total=Sum('horas'))['total'] or 0
    resumen_empleados = (
        registros.values('empleado__username')
        .annotate(total=Sum('horas'))
        .order_by('-total')
    )
    # ... (el resto del cÃƒÂ¡lculo de resumen_proyectos_con_variacion que ya tienes) ...
    proyectos_con_registros_ids = registros.values_list('proyecto_id', flat=True).distinct()
    resumen_proyectos_qs = (
        Proyecto.objects.filter(id__in=proyectos_con_registros_ids)
        .annotate(
            total_horas_registradas=Sum('registros_horas__horas'),
            horas_presupuestadas_valor=F('cantidad_h')
        )
        .order_by('-total_horas_registradas')
    )
    resumen_proyectos_con_variacion = []
    for p in resumen_proyectos_qs:
        variacion = None
        presupuestado = p.horas_presupuestadas_valor
        registrado = p.total_horas_registradas if p.total_horas_registradas is not None else 0
        if presupuestado is not None:
             try:
                 variacion = float(registrado) - float(presupuestado)
             except (ValueError, TypeError):
                 variacion = None
        resumen_proyectos_con_variacion.append({
            'nombre': p.nombre,
            'horas_presupuestadas_valor': presupuestado,
            'total_horas_registradas': registrado,
            'variacion': variacion
        })
    # ====================================================

    # ===== PASO 3: OBTÃƒâ€°N DATOS PARA SELECTS =====
    empleados = User.objects.filter(is_staff=False, is_active=True)
    proyectos = Proyecto.objects.all()
    # ===========================================

    # ===== PASO 4: CREA EL CONTEXT =====
    context = {
        'registros': registros, # Ahora 'registros' ya existe
        'empleados': empleados,
        'proyectos': proyectos,
        'total_horas': total_horas,
        'resumen_empleados': resumen_empleados,
        'resumen_proyectos': resumen_proyectos_con_variacion,
        'empleado_id': empleado_id, # 'empleado_id' tambiÃƒÂ©n debe estar definido antes
        'proyecto_id': proyecto_id, # 'proyecto_id' tambiÃƒÂ©n debe estar definido antes
    }
    # ===================================

    return render(request, 'gestion/registro_horas_admin.html', context)


# === ADMINISTRADOR: VER BITÃƒÂCORA DE ACTIVIDADES FALTA RELACIONAR EL URLS Y DEFINIR LA BITACORA===
@login_required
def ver_actividades(request):
    """Muestra la bitÃƒÂ¡cora de acciones registradas (solo admin)."""
    if not request.user.is_staff:
        return redirect('empleado_home')

    actividades = Actividad.objects.select_related('usuario').order_by('-fecha')[:100]
    return render(request, 'gestion/actividades.html', {'actividades': actividades})

#
def registrar_cliente(request):
    """
    Vista para crear un nuevo cliente.
    Solo accesible para administradores.
    """
    if not request.user.is_staff:
        return redirect('home')

    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save()
            Actividad.objects.create(usuario=request.user, accion=f"Registro el cliente '{cliente.nombre}'.")
            return redirect('admin_home')
    else:
        form = ClienteForm()

    return render(request, 'gestion/registrar_cliente.html', {'form': form})@login_required 
def reportes(request):
    """
    Muestra los 3 reportes,
    O EXPORTA A EXCEL (3 hojas)
    O EXPORTA A PDF (1 documento con 3 tablas)
    """
    
    # 1. Obtenemos la base de todos los registros (sin cambios)
    base_query = RegistroHoras.objects.select_related(
        'proyecto', 
        'empleado', 
        'proyecto__cliente'
    )
    
    # 2. Instanciamos el formulario (sin cambios)
    form = ReporteFiltroForm(request.GET)
    
    # 3. Aplicamos los filtros (sin cambios)
    if form.is_valid():
        cleaned_data = form.cleaned_data
        
        if cleaned_data.get('cliente'):
            base_query = base_query.filter(proyecto__cliente=cleaned_data.get('cliente'))
            
        if cleaned_data.get('proyecto'):
            base_query = base_query.filter(proyecto=cleaned_data.get('proyecto'))
            
        if cleaned_data.get('empleado'):
            base_query = base_query.filter(empleado=cleaned_data.get('empleado'))
            
        if cleaned_data.get('fecha_inicio'):
            base_query = base_query.filter(fecha__gte=cleaned_data.get('fecha_inicio'))
            
        if cleaned_data.get('fecha_fin'):
            base_query = base_query.filter(fecha__lte=cleaned_data.get('fecha_fin'))

    # Esta es nuestra lista principal de bitÃƒÂ¡cora
    reporte_bitacora = base_query.order_by('-fecha')

    # 4. GENERAR REPORTE POR PROYECTO (sin cambios)
    reporte_proyectos_query = reporte_bitacora.values(
        'proyecto__id',
        'proyecto__nombre',
        'proyecto__cliente__nombre',
        'proyecto__cantidad_h'
    ).annotate(
        horas_registradas_filtradas=Sum('horas')
    ).order_by('-horas_registradas_filtradas')

    reporte_proyectos = []
    for p in reporte_proyectos_query:
        presupuestadas = p['proyecto__cantidad_h']
        registradas = p['horas_registradas_filtradas']
        
        if registradas is None:
            registradas = 0
            
        if presupuestadas is not None and presupuestadas > 0:
            progreso = (registradas / presupuestadas) * 100
            horas_restantes = presupuestadas - registradas
        else:
            progreso = 0
            horas_restantes = -registradas 
            
        p['progreso'] = round(progreso, 2)
        p['horas_restantes'] = horas_restantes
        reporte_proyectos.append(p)


    # 5. GENERAR REPORTE POR EMPLEADO (sin cambios)
    reporte_empleados = reporte_bitacora.values(
        'empleado__id',
        'empleado__first_name',
        'empleado__last_name',
        'empleado__username'
    ).annotate(
        horas_totales=Sum('horas'),
        num_registros=Count('id')
    ).order_by('-horas_totales')
    
    reporte_empleados_procesado = []
    for e in reporte_empleados:
        if e['horas_totales'] is None:
            e['horas_totales'] = 0
        reporte_empleados_procesado.append(e)

    # 6. Preparamos el contexto (sin cambios)
    contexto = {
        'form': form,
        'reporte_bitacora': reporte_bitacora,
        'reporte_proyectos': reporte_proyectos,
        'reporte_empleados': reporte_empleados_procesado, 
    }

    # 
    # =======================================
    # === Ã‚Â¡LÃƒâ€œGICA DE EXPORTACIÃƒâ€œN (EXCEL Y PDF)! ===
    # =======================================
    
    export_type = request.GET.get('exportar')

    # --- OpciÃƒÂ³n 1: Exportar a EXCEL (Sin cambios) ---
    if export_type == 'excel':
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"reporte_completo_{datetime.date.today()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb = openpyxl.Workbook()
        bold_font = Font(bold=True)
        
        # Hoja 1: Resumen Proyectos
        ws1 = wb.active
        ws1.title = "Resumen Proyectos"
        headers1 = [
            "Proyecto", "Cliente", "H. Presupuestadas", 
            "H. Registradas", "H. Restantes", "Consumo (%)"
        ]
        ws1.append(headers1)
        for cell in ws1[1]: cell.font = bold_font
        for p in reporte_proyectos:
            ws1.append([
                p['proyecto__nombre'], p['proyecto__cliente__nombre'],
                p['proyecto__cantidad_h'], p['horas_registradas_filtradas'],
                p['horas_restantes'], p['progreso']
            ])

        # Hoja 2: Resumen Empleados
        ws2 = wb.create_sheet(title="Resumen Empleados")
        headers2 = ["Empleado", "Horas Totales", "NÃ‚Â° de Registros"]
        ws2.append(headers2)
        for cell in ws2[1]: cell.font = bold_font
        for e in reporte_empleados_procesado:
            full_name = f"{e['empleado__first_name']} {e['empleado__last_name']}"
            ws2.append([
                full_name.strip() or e['empleado__username'],
                e['horas_totales'], e['num_registros']
            ])

        # Hoja 3: BitÃƒÂ¡cora Detalle
        ws3 = wb.create_sheet(title="Bitacora Detalle")
        headers3 = ["Fecha", "Empleado", "Proyecto", "Cliente", "Horas", "DescripciÃƒÂ³n"]
        ws3.append(headers3)
        for cell in ws3[1]: cell.font = bold_font
        for registro in reporte_bitacora:
            ws3.append([
                registro.fecha, registro.empleado.get_full_name() or registro.empleado.username,
                registro.proyecto.nombre, registro.proyecto.cliente.nombre,
                registro.horas, registro.descripcion
            ])

        wb.save(response)
        return response

    # --- OpciÃƒÂ³n 2: Exportar a PDF (Ã‚Â¡ACTUALIZADO CON xhtml2pdf!) ---
    elif export_type == 'pdf':
        
    
        html_string = render_to_string('gestion/reporte_pdf.html', contexto)

    
        result = BytesIO()
        
    
        pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)
        
       
        if not pdf.err:
         
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            filename = f"reporte_completo_{datetime.date.today()}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        
      
        return HttpResponse(f"Error al generar el PDF: {pdf.err}", status=500)

    
    # --- OpciÃƒÂ³n 3: Mostrar la pÃƒÂ¡gina HTML normal ---
    # Si no se presionÃƒÂ³ ningÃƒÂºn botÃƒÂ³n de 'exportar', renderizamos la pÃƒÂ¡gina.
    return render(request, 'gestion/reportes.html', contexto)

def empleados(request):


  
    return render(request, 'gestion/empleados.html') #, context)


# --- LISTA DE EMPLEADOS ---
@login_required
def lista_empleados(request):
    if not request.user.is_staff:
        return redirect('empleado_home')

    empleados = (
        User.objects
        .filter(is_staff=False, is_active=True)
        .prefetch_related(
            Prefetch(
                'proyectos_como_empleado',
                queryset=Proyecto.objects.only('id', 'nombre', 'situacion')
            ),
            Prefetch(
                'asignaciones',
                queryset=AsignacionProyecto.objects.select_related('proyecto').order_by('-activo', 'proyecto__nombre')
            )
        )
        .order_by('username')
    )

    return render(request, 'gestion/empleados.html', {'empleados': empleados})


# === GESTIÃƒâ€œN DE CLIENTES (ADMIN) ===
@login_required
def lista_clientes(request):
    if not request.user.is_staff:
        return redirect('empleado_home')

    clientes = Cliente.objects.all()

    context = {
        'clientes': clientes
    }

    return render(request, 'gestion/lista_clientes.html', context)


@login_required
def editar_cliente(request, cliente_id):
    if not request.user.is_staff:
        return redirect('empleado_home')

    cliente = get_object_or_404(Cliente, id=cliente_id)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            Actividad.objects.create(
                usuario=request.user,
                accion=f"EditÃƒÂ³ el cliente '{cliente.nombre}'."
            )
            return redirect('lista_clientes')
    else:
        form = ClienteForm(instance=cliente)
    return render(request, 'gestion/editar_cliente.html', {'form': form, 'cliente': cliente})


@login_required
def eliminar_cliente(request, cliente_id):
    if not request.user.is_staff:
        return redirect('empleado_home')

    cliente = get_object_or_404(Cliente, id=cliente_id)
    if request.method == 'POST':
        nombre = cliente.nombre
        cliente.delete()
        Actividad.objects.create(
            usuario=request.user,
            accion=f"EliminÃƒÂ³ el cliente '{nombre}'."
        )
        return redirect('lista_clientes')
    return render(request, 'gestion/eliminar_cliente.html', {'cliente': cliente})


# --- ASIGNAR PROYECTO A EMPLEADO ---
@login_required
def asignar_proyecto_empleado(request, empleado_id):
    if not request.user.is_staff:
        return redirect('empleado_home')

    empleado = get_object_or_404(User, id=empleado_id, is_staff=False)

    if request.method == 'POST':
        form = AsignarProyectoForm(request.POST, empleado=empleado)
        if form.is_valid():
            asig = form.save(commit=False)
            asig.empleado = empleado
            asig.activo = True
            asig.fecha_asignacion = timezone.now().date()
            asig.save()

            Actividad.objects.create(
                usuario=request.user,
                accion=f"AsignÃƒÂ³ el proyecto '{asig.proyecto.nombre}' a {empleado.username} (rol: {asig.get_rol_en_proyecto_display()})."
            )

            return redirect('lista_empleados')
    else:
        form = AsignarProyectoForm(empleado=empleado)

    return render(request, 'gestion/asignar_proyecto.html', {
        'empleado': empleado,
        'form': form
    })


# --- DESASIGNAR (BAJA) PROYECTO DE EMPLEADO ---
@login_required
def desasignar_proyecto_empleado(request, empleado_id, proyecto_id):
    if not request.user.is_staff:
        return redirect('empleado_home')

    empleado = get_object_or_404(User, id=empleado_id, is_staff=False)
    asignacion = get_object_or_404(
        AsignacionProyecto,
        empleado=empleado,
        proyecto_id=proyecto_id,
        activo=True
    )

    if request.method == 'POST':
        asignacion.activo = False
        asignacion.fecha_baja = timezone.now().date()
        asignacion.save()

        Actividad.objects.create(
            usuario=request.user,
            accion=f"DesasignÃƒÂ³ el proyecto '{asignacion.proyecto.nombre}' de {empleado.username}."
        )

        return redirect('lista_empleados')

    # ConfirmaciÃƒÂ³n simple
    return render(request, 'gestion/confirmar_desasignacion.html', {
        'empleado': empleado,
        'proyecto': asignacion.proyecto
    })

@login_required
def registrar_cliente(request):
    """
    Vista para crear un nuevo cliente.
    Solo accesible para administradores.
    """
   
    if not request.user.is_staff:
        return redirect('home') 


    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save() 
            Actividad.objects.create(usuario=request.user, accion="Registro el cliente '" + cliente.nombre + "'.")
            return redirect('admin_home') 
    else:
        
        form = ClienteForm() 

    return render(request, 'gestion/registrar_cliente.html', {'form': form})


@login_required
def registrar_usuario(request):
    # Solo los administradores pueden registrar empleados
    if not request.user.is_staff:
        return redirect('home')

    if request.method == 'POST':
        form = EmpleadoForm(request.POST)
        if form.is_valid():
            form.save() 
            return redirect('admin_home') 
    else:
        form = EmpleadoForm()

    return render(request, 'gestion/registrar_usuario.html', {'form': form})


@login_required
def editar_usuario(request, usuario_id):
    # Solo los administradores pueden editar empleados
    if not request.user.is_staff:
        return redirect('home')

    empleado = get_object_or_404(User, id=usuario_id, is_staff=False)
    perfil = get_object_or_404(PerfilEmpleado, user=empleado)

    if request.method == 'POST':
        form = EmpleadoUpdateForm(request.POST, instance=perfil, user_instance=empleado)
        if form.is_valid():
            form.save()
            Actividad.objects.create(
                usuario=request.user,
                accion=f"EditÃƒÂ³ al usuario '{empleado.username}'."
            )
            return redirect('lista_empleados')
    else:
        form = EmpleadoUpdateForm(instance=perfil, user_instance=empleado)

    return render(request, 'gestion/editar_usuario.html', {'form': form, 'empleado': empleado})


@login_required
def eliminar_usuario(request, usuario_id):
    # Solo los administradores pueden eliminar (desactivar) empleados
    if not request.user.is_staff:
        return redirect('home')

    empleado = get_object_or_404(User, id=usuario_id, is_staff=False)

    if request.method == 'POST':
        # Desactivar usuario y dar de baja asignaciones activas
        empleado.is_active = False
        empleado.save()
        AsignacionProyecto.objects.filter(empleado=empleado, activo=True).update(
            activo=False,
            fecha_baja=timezone.now().date()
        )
        Actividad.objects.create(
            usuario=request.user,
            accion=f"DesactivÃƒÂ³ al usuario '{empleado.username}'."
        )
        return redirect('lista_empleados')

    return render(request, 'gestion/eliminar_usuario.html', {'empleado': empleado})

# === VISTA DE CAMBIO DE CONTRASEÃƒâ€˜A (RECOMENDADA) ===
class CambiarPasswordView(LoginRequiredMixin, PasswordChangeView):
    """
    Vista para que el usuario cambie su contraseÃƒÂ±a.
    Usa la vista genÃƒÂ©rica de Django para mÃƒÂ¡xima seguridad.
    """
    form_class = CustomPasswordChangeForm
    template_name = 'gestion/cambiar_password.html' # El template que crearemos en el Paso 4
    success_url = reverse_lazy('password_exitoso') # Una URL a dÃƒÂ³nde ir despuÃƒÂ©s (Paso 3)

    def get_form_kwargs(self):
        """
        Esto es clave: le pasa el usuario actual al formulario.
        """
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

# --- Vista simple para el ÃƒÂ©xito ---
# TambiÃƒÂ©n necesitarÃƒÂ¡s una vista simple que muestre un mensaje de ÃƒÂ©xito.
from django.shortcuts import render

@login_required
def password_exitoso(request):
    """Muestra un mensaje de ÃƒÂ©xito despuÃƒÂ©s de cambiar la contraseÃƒÂ±a."""
    return render(request, 'gestion/password_exitoso.html')


